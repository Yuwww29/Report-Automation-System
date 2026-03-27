import pandas as pd
import os
import math
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl import load_workbook
import importlib
importlib.reload(openpyxl)


#Reusable Functions
def autofit_columns(ws, start_col = 1, end_col = None, padding = 1):
    if end_col is None:
        end_col = ws.max_column

    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = max_length + padding

highlight_fill_grey = PatternFill(start_color = 'FFD9D9D9', end_color = 'FFD9D9D9', fill_type = 'solid')
highlight_fill_darkgrey = PatternFill(start_color = 'FFAEAAAA', end_color = 'FFAEAAAA', fill_type = 'solid')
highlight_fill_blue = PatternFill(start_color = 'FFC0E6F5', end_color = 'FFC0E6F5', fill_type = 'solid')
highlight_fill_yellow = PatternFill(start_color = 'FFFFFF00', end_color = 'FFFFFF00', fill_type = 'solid')
highlight_fill_white = PatternFill(start_color = 'FFFFFFFF', end_color = 'FFFFFFFF', fill_type = 'solid')
highlight_fill_darkblue = PatternFill(start_color = 'FF156082', end_color = 'FF156082', fill_type = 'solid')
highlight_fill_orange = PatternFill(start_color = 'E46C0A', end_color = 'E46C0A', fill_type = 'solid')
                                          
            
border1 = Border(bottom=Side(style = 'medium'))
border2 = Border(top=Side(style = 'medium'))
border3 = Border(right=Side(style = 'medium'))
border4 = Border(left=Side(style = 'medium'))
border5 = Border(right=Side(style='medium'), bottom=Side(style='medium'))
border6 = Border(right=Side(style='medium'), top=Side(style='medium'))
border7 = Border(left=Side(style='medium'), bottom=Side(style='medium'))
border8 = Border(left=Side(style='medium'),top=Side(style='medium'))
border9 = Border(top=Side(style = 'medium'), bottom=Side(style='medium'))
border10 = Border(
                    left=Side(style='medium'),
                    top=Side(style = 'medium'),
                    bottom=Side(style='medium')
                )
border11 = Border(
                    right=Side(style='medium'),
                    top=Side(style = 'medium'),
                    bottom=Side(style='medium')
                )
borderfull = Border(
                    left=Side(style='medium'),
                    right=Side(style = 'medium'),
                    top=Side(style = 'medium'),
                    bottom=Side(style='medium')
                    )

def border_range(ws, min_row, max_row, min_col, max_col, border):
                for row in ws.iter_rows(min_row = min_row, max_row = max_row,
                                        min_col = min_col, max_col = max_col):
                    for cell in row:
                        cell.border = border
            
def background_range(ws, min_row, max_row, min_col, max_col, highlight):
                for row in ws.iter_rows(min_row = min_row, max_row = max_row,
                                        min_col = min_col, max_col = max_col):
                    for cell in row:
                        cell.fill = highlight

                        if cell.value is not None and not (isinstance(cell.value, float) and math.isnan(cell.value)):
                            cell.font = Font(bold = True, size = 15)

def pivottable_range(ws, min_row, max_row, min_col, max_col, highlight, font = None, border = None):
                for row in ws.iter_rows(min_row = min_row, max_row = max_row,
                                                 min_col = min_col, max_col = max_col):
                    for cell in row:
                        cell.fill = highlight
                        cell.font = Font(size = 15)

                        if font == 1:
                            cell.font = Font(bold = True, size = 15)

                        if border == 1:
                            cell.border = Border()
                        if border == 2:
                            cell.border = Border(bottom=Side(style = 'medium', color = 'FF83CCEB'))
                        if border == 3:
                            cell.border = Border(top=Side(style = 'medium', color = 'FF83CCEB'))

def style_excel(results):
    output_file = results.get('output_file', 'output.xlsx')
    output_dir = os.path.dirname(output_file)

    os.makedirs(output_dir, exist_ok=True)

    if not os.path.exists(output_file):
        wb = openpyxl.Workbook()
        wb.save(output_file)
    
    with pd.ExcelWriter(results['output_file'], engine = 'openpyxl', mode = 'a', if_sheet_exists = 'overlay') as writer:
        TR_start_row = results['sq_pivot_table_rows'] + 3 + 5 + 15
        OAL_start_row = results['tr_pivot_table_rows'] + TR_start_row + 5
        lower_box_start = 13 + results['commando_pivot_table'].shape[0] + 4
        ref_col = results['commando_pivot_table'].shape[1]+2+4

        results['df'].to_excel(writer, sheet_name = 'Raw Data', index = False, header = True)
        results['commando_table'].to_excel(writer, sheet_name='Commandos', index = False)
        results['st_table'].to_excel(writer, sheet_name='ST', index = False)
        results['commando_Row'].to_excel(writer, sheet_name = 'Commandos', startcol= 4, index = False, header= False)
        results['st_Row'].to_excel(writer, sheet_name = 'ST', startcol = 4, index = False , header = False)
        results['date'].to_excel(writer, sheet_name = 'Flight Count', startcol = 1, startrow = 1, index = False, header = False)
        results['sq_daily_total'].to_excel(writer, sheet_name = 'Flight Count', startcol = 1, startrow = 2, index = False, header = False)
        results['tr_daily_total'].to_excel(writer, sheet_name = 'Flight Count', startcol = 1, startrow = 3, index = False, header = False)
        results['sqtr_daily_total'].to_excel(writer, sheet_name = 'Flight Count', startcol = 1, startrow = 4, index = False, header = False)
        results['date'].to_excel(writer, sheet_name = 'Flight Count', startcol = 1, startrow = 7, index = False, header = False)
        results['sqtr_daily_total'].to_excel(writer, sheet_name = 'Flight Count', startcol = 1, startrow = 8, index = False, header = False)
        results['oal_daily_total'].to_excel(writer, sheet_name = 'Flight Count', startcol = 1, startrow = 9, index = False, header = False)
        results['sqtroal_daily_total'].to_excel(writer, sheet_name = 'Flight Count', startcol = 1, startrow = 10, index = False, header = False)
        results['sq_pivot_table'].to_excel(writer, sheet_name = 'Flight Count', startcol = 0, startrow = 18, index = True, header = True)
        results['tr_pivot_table'].to_excel(writer, sheet_name = 'Flight Count', startcol = 0, startrow = TR_start_row, index = True, header = True)
        results['oal_pivot_table'].to_excel(writer, sheet_name = 'Flight Count', startcol = 0, startrow = OAL_start_row, index = True, header = True)
        results['date'].to_excel(writer, sheet_name = 'Tabulation', startcol = 2, startrow = 5, index = False, header = False)
        results['day'].to_excel(writer, sheet_name = 'Tabulation', startcol = 2, startrow = 6, index = False, header = False)
        results['commando_daily_total'].to_excel(writer, sheet_name = 'Tabulation', startcol = 2, startrow = 7, index = False, header = False)
        results['sqtr_daily_total'].to_excel(writer, sheet_name = 'Tabulation', startcol = 2, startrow = 8, index = False, header = False)
        results['percentage_daily_total'].to_excel(writer, sheet_name = 'Tabulation', startcol = 2, startrow = 9, index = False, header = False)
        results['commando_pivot_table'].to_excel(writer, sheet_name = 'Tabulation', startcol = 1, startrow = 13, index = True, header = True)
        results['date'].to_excel(writer, sheet_name = 'Tabulation', startcol = 2, startrow = lower_box_start, index = False, header = False)
        results['sq_daily_total'].to_excel(writer, sheet_name = 'Tabulation', startcol = 2, startrow = lower_box_start+1, index = False, header = False)
        results['tr_daily_total'].to_excel(writer, sheet_name = 'Tabulation', startcol = 2, startrow = lower_box_start+2, index = False, header = False)
        results['sqtr_daily_total'].to_excel(writer, sheet_name = 'Tabulation', startcol = 2, startrow = lower_box_start+3, index = False, header = False)
        results['date'].to_excel(writer, sheet_name = 'Tabulation', startcol = ref_col+1, startrow = 5, index = False, header = False)
        results['day'].to_excel(writer, sheet_name = 'Tabulation', startcol = ref_col+1, startrow = 6, index = False, header = False)
        results['st_daily_total'].to_excel(writer, sheet_name = 'Tabulation', startcol = ref_col+1, startrow = 7, index = False, header = False)
        results['st_pivot_table'].to_excel(writer, sheet_name = 'Tabulation', startcol = ref_col, startrow = 13, index = True, header = True)
        results['bay_coverage'].to_excel(writer, sheet_name = 'Bay Coverage Daily', startcol = 1, startrow = 2, index = False, header = True)
            
        def Raw_Data_Sheet():
            ws = writer.sheets['Raw Data']  
            autofit_columns(ws)
            
        def Commandos_Sheet():
            ws = writer.sheets['Commandos']

            max_row = ws.max_row
            max_col = ws.max_column

            start_cell = 'A1'
            end_cell = f'{get_column_letter(2)}{max_row}'

            ws.column_dimensions[get_column_letter(1)].width = 15
            ws.column_dimensions[get_column_letter(2)].width = 40
                
            for cells in ws[get_column_letter(1)]:
                cells.alignment = Alignment(horizontal = 'center', vertical = 'center')
            for cells in ws[get_column_letter(2)]:
                cells.alignment = Alignment(horizontal = 'center', vertical = 'center')

            tbl1_ref = f'{start_cell}:{end_cell}'
            tbl1 = Table(displayName = 'Commandos_Table', ref = tbl1_ref)

            tbl1.tableStyleInfo = TableStyleInfo(
                name = 'TableStyleMedium2',
                showFirstColumn = False,
                showLastColumn = False,
                showRowStripes = True,
                showColumnStripes = False
            )
            ws.add_table(tbl1)

            start_cell2 = 'E1'
            end_cell2 = f'{get_column_letter(max_col)}{max_row}'

            tbl2_ref = f'{start_cell2}:{end_cell2}'
            tbl2 = Table(displayName = 'Commandos_Row', ref = tbl2_ref)

            tbl2.tableStyleInfo = TableStyleInfo(
                name = 'TableStyleLight9',
                showFirstColumn = False,
                showLastColumn = False,
                showRowStripes = True,
                showColumnStripes = False
            )
            ws.add_table(tbl2)
            
            autofit_columns(ws, start_col = 5, end_col = max_col)

            highlight_fill = PatternFill(start_color ='FFFF0505',
                                            end_color = 'FFFF0505',
                                            fill_type = 'solid')

            search_string = ['PLB (R) Arrival By']
            for row in ws.iter_rows(min_row = 1, max_row = 1,
                                                min_col = 1, max_col = max_col):
                for cell in row:
                    for string in search_string:
                        if cell.value == string:
                            cell.fill = highlight_fill

            white_font = Font(color = 'FFFFFFFF')
            ws['A1'].font = white_font
            ws['B1'].font = white_font

        def ST_Sheet():
            ws = writer.sheets['ST']

            max_row = ws.max_row
            max_col = ws.max_column

            start_cell = 'A1'
            end_cell = f'{get_column_letter(2)}{max_row}'

            ws.column_dimensions[get_column_letter(1)].width = 15
            ws.column_dimensions[get_column_letter(2)].width = 40
                
            for cells in ws[get_column_letter(1)]:
                cells.alignment = Alignment(horizontal = 'center', vertical = 'center')
            for cells in ws[get_column_letter(2)]:
                cells.alignment = Alignment(horizontal = 'center', vertical = 'center')

            tbl1_ref = f'{start_cell}:{end_cell}'
            tbl1 = Table(displayName = 'ST_Table', ref = tbl1_ref)

            tbl1.tableStyleInfo = TableStyleInfo(
                name = 'TableStyleMedium2',
                showFirstColumn = False,
                showLastColumn = False,
                showRowStripes = True,
                showColumnStripes = False
            )
            ws.add_table(tbl1)

            start_cell2 = 'E1'
            end_cell2 = f'{get_column_letter(max_col)}{max_row}'

            tbl2_ref = f'{start_cell2}:{end_cell2}'
            tbl2 = Table(displayName = 'ST_Row', ref = tbl2_ref)

            tbl2.tableStyleInfo = TableStyleInfo(
                name = 'TableStyleLight8',
                showFirstColumn = False,
                showLastColumn = False,
                showRowStripes = True,
                showColumnStripes = False
            )
            ws.add_table(tbl2)
            
            autofit_columns(ws, start_col = 5, end_col = max_col)

            highlight_fill = PatternFill(start_color = 'FFFF0505',
                                            end_color = 'FFFF0505',
                                            fill_type = 'solid')

            search_string = ['MAB_BY', 'PLB Docking By', 'PLB (R) Arrival By', 'Pax step docking by']
            for row in ws.iter_rows(min_row = 1, max_row = 1,
                                        min_col = 1, max_col = max_col):
                for cell in row:
                    for string in search_string:
                        if cell.value == string:
                            cell.fill = highlight_fill

            white_font = Font(color = 'FFFFFFFF')
            ws['A1'].font = white_font
            ws['B1'].font = white_font

        Raw_Data_Sheet()
        Commandos_Sheet()
        ST_Sheet()
        
        def Flight_Count():           
            ws = writer.sheets['Flight Count']

            max_row = ws.max_row
            max_col = ws.max_column

            SQ_start_row = 19
            SQ_end_row = 19+results['sq_pivot_table_rows']

            tr_start_row = results['sq_pivot_table_rows'] + 3 + 5 + 16
            tr_end_row = tr_start_row + results['tr_pivot_table_rows']
            oal_start_row = results['tr_pivot_table_rows'] + tr_start_row + 5
            oal_end_row = oal_start_row + results['oal_pivot_table_rows']

            #For Flight Header 1
            ws['A2'].value = 'Date'
            ws['A3'].value = 'SQ Arrival Flights'
            ws['A4'].value = 'TR Arrival Flights'
            ws['A5'].value = 'SQ & TR Arrival Flights'

            ws['A8'].value = 'Date'
            ws['A9'].value = 'SQ & TR Arrival Flights'
            ws['A10'].value = 'OAL Arrival Flights'
            ws['A11'].value = 'Grand Total'

            ws[f'{get_column_letter(max_col)}2'].value = 'Grand Total'
            ws[f'{get_column_letter(max_col)}8'].value = 'Grand Total'


            #Background
            background_range(ws, 2, 5, 1, max_col, highlight_fill_grey)
            background_range(ws, 8, 11, 1, max_col, highlight_fill_blue)
            background_range(ws, 3 ,5 ,max_col, max_col, highlight_fill_yellow)
            background_range(ws, 9, 11, max_col, max_col, highlight_fill_yellow)

            #Upper Header 
            border_range(ws, 2, 2, 1, max_col, border1)
            border_range(ws, 5, 5, 1, max_col ,border2)
            border_range(ws, 3, 4, 1, 1, border3)
            border_range(ws, 3, 4, max_col, max_col, border4)
            border_range(ws, 2, 2, 1, 1, border5)
            border_range(ws, 5, 5, 1, 1, border6)
            border_range(ws, 2, 2, max_col, max_col, border7)
            border_range(ws, 5, 5, max_col, max_col, border8)

            #Lower Header
            border_range(ws, 8, 8, 1, max_col, border1)
            border_range(ws, 11, 11, 1, max_col ,border2)
            border_range(ws, 9, 10, 1, 1, border3)
            border_range(ws, 9, 10, max_col, max_col, border4)
            border_range(ws, 8, 8, 1, 1, border5)
            border_range(ws, 11, 11, 1, 1, border6)
            border_range(ws, 8, 8, max_col, max_col, border7)
            border_range(ws, 11, 11, max_col, max_col, border8)


            pivottable_range(ws, SQ_start_row, SQ_end_row, 1, max_col, highlight_fill_white, None, 1)
            pivottable_range(ws, SQ_start_row, SQ_start_row, 1, max_col, highlight_fill_blue, 1, 2)
            pivottable_range(ws, SQ_end_row, SQ_end_row, 1, max_col, highlight_fill_blue, 1, 3)

            pivottable_range(ws, tr_start_row, tr_end_row, 1, max_col, highlight_fill_white, None, 1)
            pivottable_range(ws, tr_start_row, tr_start_row, 1, max_col, highlight_fill_blue, 1, 2)
            pivottable_range(ws, tr_end_row, tr_end_row, 1, max_col, highlight_fill_blue, 1, 3)

            pivottable_range(ws, oal_start_row, oal_end_row, 1, max_col, highlight_fill_white, None, 1)
            pivottable_range(ws, oal_start_row, oal_start_row, 1, max_col, highlight_fill_blue, 1, 2)
            pivottable_range(ws, oal_end_row, oal_end_row, 1, max_col, highlight_fill_blue, 1, 3)

            #Formatting Cells
            ws.column_dimensions['A'].width = 30
            for col in range(2, max_col+1):
                ws.column_dimensions[get_column_letter(col)].width = 18
            
            for row in ws.iter_rows(min_row = 1, max_row = max_row,
                                    min_col = 2, max_col = max_col):
                for cell in row:
                    cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

            for row in ws.iter_rows(min_row = 1, max_row = max_row,
                                    min_col = 1, max_col = 1):
                for cell in row:
                    cell.alignment = Alignment(horizontal = 'left', vertical = 'center')

            #Headers
            ws.merge_cells(f'A17:{get_column_letter(max_col)}18')
            ws['A17'] = 'SQ Arrival WB Pax Flights'
            ws['A17'].font = Font(size = 20)
            ws['A17'].alignment = Alignment(horizontal = 'center', vertical = 'center')

            ws.merge_cells(f'A{tr_start_row-2}:{get_column_letter(max_col)}{tr_start_row-1}')
            ws[f'A{tr_start_row-2}'] = 'TR Arrival WB Pax Flights'
            ws[f'A{tr_start_row-2}'].font = Font(size = 20)
            ws[f'A{tr_start_row-2}'].alignment = Alignment(horizontal = 'center', vertical = 'center')

            ws.merge_cells(f'A{oal_start_row-2}:{get_column_letter(max_col)}{oal_start_row-1}')

            ws[f'A{oal_start_row-2}'] = 'OAL Arrival WB Pax Flights'
            ws[f'A{oal_start_row-2}'].font = Font(size = 20)
            ws[f'A{oal_start_row-2}'].alignment = Alignment(horizontal = 'center', vertical = 'center')

            border_range(ws, 17, 18, 1, max_col, borderfull)
            border_range(ws, tr_start_row-2, tr_start_row-1, 1, max_col, borderfull)
            border_range(ws, oal_start_row-2, oal_start_row-1, 1, max_col, borderfull)

        Flight_Count()
            
        def Commando_Tabulation():
            commando_pivot_table_start_row = 14
            commando_pivot_table_end_row = 14 + results['commando_pivot_table'].shape[0]

            ws = writer.sheets['Tabulation']
            ws.sheet_view.showGridLines = False

            max_col = results['commando_pivot_table'].shape[1]+2

            #Manual Write in

            ws['B6'].value = 'Commandos'
            ws['B8'].value = 'Commandos'
            ws['B9'].value = 'SQ & TR Flights (WB, ARRIVAL, PAX)'
            ws['B10'].value = 'Percentage'
            ws[f'{get_column_letter(max_col)}6'].value = 'Total'

            ws[f'B{lower_box_start+2}'].value = 'SQ WB Arrival Pax Flights'
            ws[f'B{lower_box_start+3}'].value = 'TR WB Arrival Pax Flights'
            ws[f'B{lower_box_start+4}'].value = 'Grand Total'
            
            ws[f'{get_column_letter(max_col)}{lower_box_start+1}'].value = 'Total'
            
            background_range(ws, 6, 10 , 2, max_col, highlight_fill_blue)
            background_range(ws, 6, 6, 2, 2, highlight_fill_darkgrey)
            background_range(ws, 8, 10, max_col, max_col, highlight_fill_yellow)
            background_range(ws, lower_box_start+1, lower_box_start+4, 2, max_col, highlight_fill_blue)
            background_range(ws, lower_box_start+2, lower_box_start+4, max_col, max_col, highlight_fill_yellow)


            border_range(ws, 8, 8, 3, max_col-1, border2)
            border_range(ws, 9, 9, 3, max_col-1, border1)
            border_range(ws, 10, 10, 3, max_col-1, border1)
            border_range(ws, 8, 8, 2, 2, border8)
            border_range(ws, 9, 9, 2, 2, border7)
            border_range(ws, 10, 10, 2, 2, border7)
            border_range(ws, 8, 8, max_col, max_col, border6)
            border_range(ws, 9, 9, max_col, max_col, border5)
            border_range(ws, 10, 10, max_col, max_col, border5)
            border_range(ws, lower_box_start+2, lower_box_start+4, 2, max_col, borderfull)


            pivottable_range(ws, commando_pivot_table_start_row+1, commando_pivot_table_end_row-1, 2, max_col, highlight_fill_white, None, 1)
            pivottable_range(ws, commando_pivot_table_start_row, commando_pivot_table_start_row, 2, max_col, highlight_fill_blue, 1, 2)
            pivottable_range(ws, commando_pivot_table_end_row, commando_pivot_table_end_row, 2, max_col, highlight_fill_blue, 1, 3)
            pivottable_range(ws, commando_pivot_table_start_row+1, commando_pivot_table_end_row, max_col, max_col, highlight_fill_yellow)

            for row in ws.iter_rows(min_row = commando_pivot_table_end_row, max_row = commando_pivot_table_end_row, min_col = 2, max_col = max_col):
                    for cell in row:
                        cell.font = Font(size = 15,
                                        bold = True)
                        
            ws.column_dimensions['B'].width = 50
            for col in range(3, max_col+1):
                ws.column_dimensions[get_column_letter(col)].width = 18
            
            for row in ws.iter_rows(min_row = 1, max_row = lower_box_start+4,
                                    min_col = 3, max_col = max_col+1):
                for cell in row:
                    cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

            for row in ws.iter_rows(min_row = 1, max_row = lower_box_start+4,
                                    min_col = 2, max_col = 2):
                for cell in row:
                    cell.alignment = Alignment(horizontal = 'left', vertical = 'center')


            ws.merge_cells(f'B5:{get_column_letter(max_col)}5')
            ws['B5'] = 'Output = PLB (R) Arrival, filtering ARRIVAL, SQ & TR, WB Flights'
            ws['B5'].fill = highlight_fill_yellow
            ws['B5'].font = Font(size = 12)
            ws['B5'].alignment = Alignment(horizontal = 'center', vertical = 'center')
            for i in range(3,max_col+1):
                    ws[f'{get_column_letter(i)}10'].number_format = '0.00%'

        Commando_Tabulation()
        
        def ST_Tabulation():
            ws = writer.sheets['Tabulation']
            ws.sheet_view.showGridLines = False

            start_col = results['commando_pivot_table'].shape[1]+2+5
            last_col = results['commando_pivot_table'].shape[1]+2+5+results['st_pivot_table'].shape[1]

            ST_start_row = 14
            ST_end_row = 14+results['st_pivot_table'].shape[0]

            #Manual Write In
            ws[f'{get_column_letter(start_col)}6'].value = 'ST'
            ws[f'{get_column_letter(start_col)}8'].value = 'ST'
            ws[f'{get_column_letter(last_col)}6'].value = 'Total'

            background_range(ws, 6, 8, start_col, last_col, highlight_fill_blue)
            background_range(ws, 6, 6, start_col, start_col, highlight_fill_darkgrey)
            background_range(ws, 8, 8, last_col, last_col, highlight_fill_yellow)

            border_range(ws, 8, 8, start_col, start_col, border10)
            border_range(ws, 8, 8, last_col, last_col, border11)
            border_range(ws, 8, 8 , start_col+1, last_col-1, border9)

            pivottable_range(ws, ST_start_row+1, ST_end_row-1, start_col, last_col, highlight_fill_white, None, 1)
            pivottable_range(ws, ST_start_row, ST_start_row, start_col, last_col, highlight_fill_blue, 1, 2)
            pivottable_range(ws, ST_end_row, ST_end_row, start_col, last_col, highlight_fill_blue, 1, 3)
            pivottable_range(ws, ST_start_row+1, ST_end_row, last_col, last_col, highlight_fill_yellow)

            for row in ws.iter_rows(min_row = ST_end_row, max_row = ST_end_row, min_col = last_col, max_col = last_col):
                    for cell in row:
                        cell.font = Font(size = 15,
                                        bold = True)

            ws.column_dimensions[f'{get_column_letter(start_col)}'].width = 50
            for col in range(start_col+1, last_col+1):
                    ws.column_dimensions[get_column_letter(col)].width = 18

            for row in ws.iter_rows(min_row = 1, max_row = ST_end_row,
                                    min_col = start_col, max_col = last_col):
                for cell in row:
                    cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

            for row in ws.iter_rows(min_row = 1, max_row = ST_end_row,
                                    min_col = start_col, max_col = start_col):
                for cell in row:
                    cell.alignment = Alignment(horizontal = 'left', vertical = 'center')

            ws.merge_cells(f'{get_column_letter(start_col)}5:{get_column_letter(last_col)}5')
            ws[f'{get_column_letter(start_col)}5'] = 'Output = MAB, PLB Docking End, PLB (R) Arrival & Pax Step Docking' 
            ws[f'{get_column_letter(start_col)}5'].fill = highlight_fill_yellow
            ws[f'{get_column_letter(start_col)}5'].font = Font(size = 12)
            ws[f'{get_column_letter(start_col)}5'].alignment = Alignment(horizontal = 'center', vertical = 'center')

        ST_Tabulation()

        def Bay_Coverage_Daily():
            ws = writer.sheets['Bay Coverage Daily']

            max_row = ws.max_row
            max_col = ws.max_column
            
            border_range(ws, 3, max_row, 2, max_col, borderfull)
            for row in ws.iter_rows(min_row = 3, max_row = max_row,
                                    min_col = 2, max_col = max_col):
                 for cell in row:
                      cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
                      cell.font = Font(size = 15,
                                       bold = True)
            ws.merge_cells(f'B2:{get_column_letter(max_col)}2')
            ws['B3'] = 'Dates'
            ws['B2'] = 'Coverage of Bays - Flights Docked by Commandos'
            ws['B2'].fill = highlight_fill_yellow
            ws['B2'].font = Font(size = 15)
            ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(2, max_col+1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            for row in range(3, max_row+1):
                ws.row_dimensions[row].height = 25
        
        Bay_Coverage_Daily()

    wb = load_workbook(results['output_file'])
    wb._sheets[0], wb._sheets[5] = wb._sheets[5], wb._sheets[0]
    wb._sheets[1], wb._sheets[4] = wb._sheets[4], wb._sheets[1]
    
    wb.remove(wb['Sheet'])

    wb._sheets[4], wb._sheets[5] = wb._sheets[5], wb._sheets[4]
    
    wb.save(results['output_file'])
    output_file = results['output_file']

    print('Styling Done!')

    return output_file
